"""
Generador de Códigos de Barras - OTROS EQUIPOS
===============================================
- Conecta a SQL Server y obtiene datos de Otros_Equipos
- Genera código de barras Code128 con: Empresa | Edificio | Area | Tipo | #00
- Medidas: 2.0 cm × 1.0 cm por etiqueta
- Agrupa en hojas tamaño carta para imprimir y recortar
"""

import pyodbc
import os
from collections import defaultdict
from openpyxl import load_workbook
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.units import cm, mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.graphics.barcode import code128

# ─────────────────────────────────────────────
#  CONFIGURACIÓN
# ─────────────────────────────────────────────
SQL_CONFIG = {
    "server": "150.1.1.152",
    "database": "BI",
    "username": "ConsultaBD",
    "password": "5D$bc#kM&5W2T8J40?s%",
    "driver": "{ODBC Driver 17 for SQL Server}",
}

TABLA_SQL = "[BI].[Inventario].[Otros_Equipos]"

OUTPUT_DIR = "barcode_otros_output"
PDF_SALIDA = os.path.join(OUTPUT_DIR, "barcode_otros_equipos.pdf")
EXCEL_INVENTARIO = "inventario.xlsx"
HOJA_AREAS = "Areas"
HOJA_MATERIALES = "Materiales"

# ── Medidas de etiqueta ────────────────────────────
BAR_W_CM   = 4.0          # Ancho del código de barras
BAR_H_CM   = 2.0          # Alto del código de barras
LBL_W_CM   = 5.0          # Ancho total de la etiqueta (margen lateral)
LBL_H_CM   = 2.5          # Alto total (barras + texto)
MARGIN_CM  = 0.5          # Márgenes de página
GAP_CM     = 0.5         # Espacio entre etiquetas


# ─────────────────────────────────────────────
#  CONEXIÓN A BASE DE DATOS
# ─────────────────────────────────────────────
def get_connection():
    conn_str = (
        f"DRIVER={SQL_CONFIG['driver']};"
        f"SERVER={SQL_CONFIG['server']};"
        f"DATABASE={SQL_CONFIG['database']};"
        f"UID={SQL_CONFIG['username']};"
        f"PWD={SQL_CONFIG['password']};"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str, timeout=10)


def query_to_dicts(cursor, sql):
    """Ejecuta SQL y retorna lista de diccionarios."""
    cursor.execute(sql)
    cols = [d[0] for d in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


# ─────────────────────────────────────────────
#  OBTENER DATOS
# ─────────────────────────────────────────────
def obtener_otros_equipos(conn):
    """Obtiene solo los campos requeridos de la tabla Otros_Equipos."""
    cur = conn.cursor()
    sql = f"""
        SELECT
            [Tipo] AS Tipo,
            [Empresa] AS Empresa,
            [edificio] AS Edificio,
            [Area] AS Area
        FROM {TABLA_SQL}
    """
    return query_to_dicts(cur, sql)


EMPRESAS_ABREVIADAS = {
    "ABSOLUTE BROKERAGE CUSTOMS, S.C.": "ABC",
}

EMPRESAS_POR_CONTENIDO = {
    "SUPER EXPRESS DE CARGA FEDERAL, S.A DE C.V": "SEC",
}

def limpiar_valor(valor) -> str:
    if valor is None:
        return "-"
    texto = str(valor).strip().upper()
    if not texto:
        return "-"
    texto = texto.replace("|", "/")
    return " ".join(texto.split())


def token_codigo(valor, longitud: int) -> str:
    valor = limpiar_valor(valor)
    if valor == "-":
        return "X" * longitud

    solo_validos = "".join(ch for ch in valor if ch.isalnum())
    if not solo_validos:
        return "X" * longitud
    return solo_validos[:longitud].ljust(longitud, "X")


def abreviar_texto(valor: str, longitud: int) -> str:
    valor = limpiar_valor(valor)
    if valor == "-":
        return valor
    return valor[:longitud]


def cargar_acronimos_areas() -> dict:
    """Carga el catalogo Area -> Acronimo desde inventario.xlsx."""
    wb = load_workbook(EXCEL_INVENTARIO, read_only=True, data_only=True)
    try:
        ws = wb[HOJA_AREAS]
        rows = ws.iter_rows(min_row=1, values_only=True)
        encabezados = [str(col).strip() if col is not None else "" for col in next(rows)]
        idx_area = encabezados.index("Area")
        idx_acronimo = encabezados.index("Acronimo")

        acronimos = {}
        for row in rows:
            area = limpiar_valor(row[idx_area] if idx_area < len(row) else None)
            acronimo = limpiar_valor(row[idx_acronimo] if idx_acronimo < len(row) else None)
            if area != "-" and acronimo != "-":
                acronimos[area] = acronimo
        return acronimos
    finally:
        wb.close()


def cargar_abreviaturas_materiales() -> dict:
    """Carga el catalogo Material -> Abreviatura desde inventario.xlsx."""
    wb = load_workbook(EXCEL_INVENTARIO, read_only=True, data_only=True)
    try:
        ws = wb[HOJA_MATERIALES]
        rows = ws.iter_rows(min_row=1, values_only=True)
        encabezados = [str(col).strip() if col is not None else "" for col in next(rows)]
        idx_material = encabezados.index("Material")
        idx_abreviatura = encabezados.index("Abreviatura")

        abreviaturas = {}
        for row in rows:
            material = limpiar_valor(row[idx_material] if idx_material < len(row) else None)
            abreviatura = limpiar_valor(row[idx_abreviatura] if idx_abreviatura < len(row) else None)
            if material != "-" and abreviatura != "-":
                abreviaturas[material] = abreviatura
        return abreviaturas
    finally:
        wb.close()


def abreviar_empresa(valor) -> str:
    texto = limpiar_valor(valor)
    if texto in EMPRESAS_ABREVIADAS:
        return EMPRESAS_ABREVIADAS[texto]
    for patron, abreviatura in EMPRESAS_POR_CONTENIDO.items():
        if patron in texto:
            return abreviatura
    if texto == "-":
        return texto
    return token_codigo(texto, 3)


def abreviar_area(valor, acronimos_areas: dict) -> str:
    texto = limpiar_valor(valor)
    if texto in acronimos_areas:
        return acronimos_areas[texto]
    if texto == "-":
        return texto
    return token_codigo(texto, 3)


def abreviar_edificio(valor) -> str:
    texto = limpiar_valor(valor)
    if texto == "NORTE 180":
        return "N180"
    if texto == "NORTE 182":
        return "N182"
    if texto == "-":
        return texto
    return token_codigo(texto, 5)


def abreviar_tipo(valor, abreviaturas_materiales: dict) -> str:
    texto = limpiar_valor(valor)
    if texto in abreviaturas_materiales:
        return abreviaturas_materiales[texto]
    if texto == "-":
        return texto
    return token_codigo(texto, 4)


def asignar_consecutivo_por_area_y_tipo(equipos: list, abreviaturas_materiales: dict) -> list:
    """
    Asigna un consecutivo por combinacion Area + Tipo para identificar cada
    mobiliario dentro de la misma ubicacion.
    """
    contadores = defaultdict(int)

    for equipo in equipos:
        area_normalizada = limpiar_valor(equipo.get("Area"))
        tipo_abreviado = abreviar_tipo(equipo.get("Tipo"), abreviaturas_materiales)
        clave_conteo = (area_normalizada, tipo_abreviado)
        contadores[clave_conteo] += 1
        equipo["ConsecutivoArea"] = contadores[clave_conteo]

    return equipos


def construir_codigo_barra(equipo: dict, acronimos_areas: dict, abreviaturas_materiales: dict) -> str:
    """
    Construye un contenido corto y estable para Code128:
    Empresa|Edificio|Area|Tipo|#00
    """
    empresa = abreviar_empresa(equipo.get("Empresa"))
    edificio = abreviar_edificio(equipo.get("Edificio"))
    area = abreviar_area(equipo.get("Area"), acronimos_areas)
    tipo = abreviar_tipo(equipo.get("Tipo"), abreviaturas_materiales)
    consecutivo = str(equipo.get("ConsecutivoArea", 0)).zfill(2)
    return f"{empresa}|{edificio}|{area}|{tipo}|#{consecutivo}"


def construir_texto_etiqueta(equipo: dict, acronimos_areas: dict, abreviaturas_materiales: dict) -> str:
    empresa = abreviar_empresa(equipo.get("Empresa"))
    edificio = abreviar_edificio(equipo.get("Edificio"))
    area = abreviar_area(equipo.get("Area"), acronimos_areas)
    tipo = abreviar_tipo(equipo.get("Tipo"), abreviaturas_materiales)
    consecutivo = str(equipo.get("ConsecutivoArea", 0)).zfill(2)
    return f"{empresa} | {edificio} | {area} | {tipo} | #{consecutivo}"


def dibujar_guias_corte(c, x, y, lbl_w, lbl_h):
    """Dibuja el borde gris y las líneas de corte en las 4 esquinas."""
    cut = 3 * mm

    # Borde de etiqueta
    c.setStrokeColor(colors.HexColor("#BBBBBB"))
    c.setLineWidth(0.4)
    c.rect(x, y, lbl_w, lbl_h, fill=0, stroke=1)

    # Marcas de esquina
    c.setStrokeColor(colors.HexColor("#999999"))
    c.setLineWidth(0.25)
    for sx, sy in [(x, y + lbl_h), (x + lbl_w, y + lbl_h),
                   (x, y),         (x + lbl_w, y)]:
        dx = cut if sx == x else -cut
        c.line(sx + dx, sy, sx, sy)
        dy = -cut if sy == y + lbl_h else cut
        c.line(sx, sy + dy, sx, sy)


# ─────────────────────────────────────────────
#  GENERAR PDF
# ─────────────────────────────────────────────
def generar_pdf(equipos: list):
    acronimos_areas = cargar_acronimos_areas()
    abreviaturas_materiales = cargar_abreviaturas_materiales()
    equipos = asignar_consecutivo_por_area_y_tipo(equipos, abreviaturas_materiales)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    page_w, page_h = LETTER  # Tamaño en puntos
    margin = MARGIN_CM * cm
    gap    = GAP_CM    * cm
    lbl_w  = LBL_W_CM  * cm
    lbl_h  = LBL_H_CM  * cm
    bar_w  = BAR_W_CM  * cm
    bar_h  = BAR_H_CM  * cm

    # Calcular layout
    cols = int((page_w - 2 * margin + gap) / (lbl_w + gap))
    rows = int((page_h - 2 * margin + gap) / (lbl_h + gap))
    per_page = cols * rows

    c = canvas.Canvas(PDF_SALIDA, pagesize=LETTER)

    total = len(equipos)
    print(f"\n  Total de equipos:  {total}")
    print(f"  Etiquetas/hoja:    {per_page}  ({cols} col × {rows} fil)")
    print(f"  Páginas:           {-(-total // per_page) if total else 0}\n")

    for idx, equipo in enumerate(equipos):
        page_idx = idx % per_page

        if page_idx == 0 and idx > 0:
            c.showPage()

        col = page_idx % cols
        row = page_idx // cols

        x = margin + col * (lbl_w + gap)
        y = page_h - margin - (row + 1) * lbl_h - row * gap

        # Guías de corte
        dibujar_guias_corte(c, x, y, lbl_w, lbl_h)

        # ── Generar código de barras ───────────────────────────
        codigo = construir_codigo_barra(equipo, acronimos_areas, abreviaturas_materiales)
        texto_etiqueta = construir_texto_etiqueta(equipo, acronimos_areas, abreviaturas_materiales)
        barcode = code128.Code128(
            codigo,
            barHeight=bar_h * 0.68,
            barWidth=0.20 * mm,
            humanReadable=False,
        )

        # Centrar barcode horizontalmente
        bar_x = x + (lbl_w - barcode.width) / 2
        bar_y = y + lbl_h - bar_h - (1.5 * mm)
        barcode.drawOn(c, bar_x, bar_y + 3 * mm)

        # ── Información textual debajo ────────────────────────
        c.setStrokeColor(colors.HexColor("#DDDDDD"))
        c.setLineWidth(0.3)
        c.line(x + 2*mm, bar_y - 1*mm, x + lbl_w - 2*mm, bar_y - 1*mm)

        c.setFont("Helvetica", 4.5)
        c.setFillColor(colors.HexColor("#888888"))
        c.drawCentredString(x + lbl_w / 2, bar_y - 4*mm, "OTROS EQUIPOS")

        # Mostrar en formato compacto: Empresa | Edif | Area | Tipo | #00
        c.setFont("Helvetica", 5)
        c.setFillColor(colors.black)
        c.drawCentredString(x + lbl_w / 2, bar_y - 7*mm, texto_etiqueta)

        if (idx + 1) % 10 == 0 or idx + 1 == total:
            print(f"  ✓ Procesados {idx + 1}/{total} equipos...", end="\r")

    c.save()
    print(f"\n\n  ✅ PDF generado: {PDF_SALIDA}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  GENERADOR DE CÓDIGOS DE BARRAS - OTROS EQUIPOS")
    print("  Formato: Empresa | Edificio | Area | #00")
    print("  Medida:  2.0 × 1.0 cm por etiqueta")
    print("=" * 60)

    print("\n[1/3] Conectando a SQL Server...")
    try:
        conn = get_connection()
        print("  ✅ Conexión exitosa")
    except Exception as e:
        print(f"  ❌ Error de conexión: {e}")
        return

    print("\n[2/3] Obteniendo equipos de Otros_Equipos...")
    try:
        equipos = obtener_otros_equipos(conn)
        print(f"  ✅ {len(equipos)} equipos encontrados")
    except Exception as e:
        print(f"  ❌ Error leyendo tabla: {e}")
        conn.close()
        return

    conn.close()

    if not equipos:
        print("\n  ⚠ No se encontraron equipos en la tabla.")
        return

    print("\n[3/3] Generando PDF con códigos de barras...")
    generar_pdf(equipos)

    print("\n" + "=" * 60)
    print(f"  Archivo listo: {PDF_SALIDA}")
    print("  Imprime en tamaño carta, sin escalar (100%)")
    print("  Cada etiqueta mide 2.0 × 1.0 cm")
    print("=" * 60)


if __name__ == "__main__":
    main()
